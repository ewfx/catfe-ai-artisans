from flask import Flask, request, render_template, send_file, jsonify, Response
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import google.generativeai as genai
import os
import pandas as pd
import subprocess

app = Flask(__name__)

def load_config():
    try:
        file_path = os.path.join(script_dir, "config.json")
        with open(file_path, "r") as config_file:
            return json.load(config_file)
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})  

script_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(script_dir, "Generated/Test_Scenarios_Test_Data.xlsx")
upload_file_path = os.path.join(script_dir, "TestFiles")
test_file_path = os.path.join(script_dir, "TestFiles/Test_Script.py")
result_file_path = os.path.join(script_dir, "Generated/Actual_Test_Results.txt")
test_cases = None
CONFIG = load_config()
API_KEY = CONFIG["gemini_api_key"]
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel("gemini-2.0-flash")

def load_instructions():
    try:
        file_path = os.path.join(script_dir, "instructions.md")
        with open(file_path, "r") as instructions_file:
            return instructions_file.read()
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})  

def generate_tests_prompt(context, num_cases=10):
    instructions = load_instructions()
    return f"""
    You are an expert in financial transactions and risk assessment. Generate {num_cases} test cases for the following scenario:
    
    Scenario: {context}
    
    Based on the context, follow these instructions:
    {instructions}

    Make sure to add the edge cases
    
    Each test case should include:
    - Test Case ID
    - Test scenario (gherkin format, do not use '<br>')
    - Test data (use '<br>' for line breaks)
    - Validation steps (use '<br>' for line breaks)
    - Expected Results (use '<br>' for line breaks)
    
    Validate the test cases for accuracy, completeness, and relevance.
    
    Format the output as a table with the following columns:
    | Test Case ID | Test Scenario | Test Data | Validation Steps | Expected Results
    """

def generate_tests_prompt_with_code(context, num_cases, upload_file_path):
    instructions = load_instructions()
    with open(upload_file_path, "r") as f:
        python_code = f.read()
    return f"""
    You are an expert in financial transactions and risk assessment. Generate {num_cases} test cases for the following code which is based on the scenario:
    
    Scenario: {context}

    Code: {python_code}
    
    Based on the context, follow these instructions:
    {instructions}

    Make sure to add the edge cases
    
    Each test case should include:
    - Test Case ID
    - Test scenario (gherkin format, do not use '<br>')
    - Test data (use '<br>' for line breaks)
    - Validation steps (use '<br>' for line breaks)
    - Expected Results (use '<br>' for line breaks)
    
    Validate the test cases for accuracy, completeness, and relevance.
    
    Format the output as a table with the following columns:
    | Test Case ID | Test Scenario | Test Data | Validation Steps | Expected Results

    Do not add any additional text to the response
    """

def generate_test_cases(context, num_cases=10, upload_file_path=None, code_uploaded=False):
    if code_uploaded:
        prompt = generate_tests_prompt_with_code(context, num_cases, upload_file_path)
    else:
        prompt = generate_tests_prompt(context, num_cases)

    global test_cases
    
    try:
        response = model.generate_content(prompt)
        test_cases = response.text.strip()
        if (code_uploaded==False):
            save_to_excel(test_cases, context)
        return jsonify({"message": "Test cases generated successfully."})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})  

def save_to_excel(test_cases, context, executed=False):
    rows = []
    try:
        test_cases = test_cases.replace("**", "")
        lines = test_cases.split("\n")
        if executed:
            headers = ["Test Case ID", "Test Scenario", "Test Data", "Validation Steps", "Expected Results", "Test Case Results"]
            start = 2
        else:
            headers = ["Test Case ID", "BDD Format test case", "Test Data", "Validation Steps", "Expected Results"]
            start = 3
        


        for line in lines[start:]:
            if line.strip(): 
                columns = [cell.strip() for cell in line.split("|")[1:-1]]
                if(len(columns) > 1):
                    for i in range(0, 5): 
                        if "<br>" in columns[i]: 
                            lines_in_column = columns[i].split("<br>")
                            numbered_lines = [
                                f"{idx + 1}. {line.strip()}" if not line.strip().startswith(f"{idx + 1}.") else line.strip()
                                for idx, line in enumerate(lines_in_column) if line.strip()
                            ]
                            columns[i] = "\n".join(numbered_lines)
                    rows.append(columns)

        df = pd.DataFrame(rows, columns=headers)
        df.to_excel(excel_file_path, index=False, engine='openpyxl')
        
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for cell in sheet[1]:  
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if executed:
            table_range = f"A1:F{len(rows) + 1}"
        else:
            table_range = f"A1:E{len(rows) + 1}" 
        table = Table(displayName="TestCasesTable", ref=table_range)

        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        workbook.save(excel_file_path)
        return jsonify({"message": "Test cases saved successfully."})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})  

def generate_pytest_tests(python_code, filename):
    prompt = f"""Generate pytest test cases for the following Python code:\n\n{python_code}\n
      
      Use the {test_cases} as cases for the corresponding pytest. Use the Test Case ID as the test method name\n
      
      Use {filename} as my module\n 
      
      Please give only the python code and no other text"""
    response = model.generate_content(prompt)
    return response.text if response else "Failed to generate test cases."

def automate_python_testing(py_file, filename):
    try:
        with open(py_file, "r") as f:
            python_code = f.read()
        test_code = generate_pytest_tests(python_code, filename)
        test_file = os.path.join(script_dir, "TestFiles/Test_Script.py")
        with open(test_file, "w") as f:
            lines = test_code.splitlines() 
            trimmed_code = "\n".join(lines[1:-1]) 
            f.write(trimmed_code) 

        with open(result_file_path, "w") as result_file:
            result_file.write("\n\n Running tests with coverage analysis...\n")
            coverage_path = os.path.join(script_dir, "TestFiles/.")
            subprocess.run(["coverage", "run", str(f"--source={coverage_path}"), "-m", "pytest", test_file], stdout=result_file, stderr=result_file)
            subprocess.run(["coverage", "report"], stdout=result_file, stderr=result_file)
            subprocess.run(["coverage", "html"], stdout=result_file, stderr=result_file)
            result_file.write("\n\n Test execution and coverage report completed!\n")
        
        return jsonify({"message": "Test cases generated and executed successfully."})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})  

def save_test_results(context):
    global test_cases
    with open(result_file_path, "r") as result_file:
        results = result_file.read()  
    prompt = f""" In the {test_cases} add an additional column | Actual Results.
    
    The Actual result should contain Pass or Fail depending on the failures mentioned in {results}

    Modify the {test_cases} to output as following :
    | Test Case ID | Test Scenario | Test Data | Validation Steps | Expected Results | Actual Results

    Do not add any additional text to the response
    """
    try:
        response = model.generate_content(prompt)
        test_cases = response.text.strip()
        save_to_excel(test_cases, context, True)
        return jsonify({"message": "Test cases generated and executed successfully."})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})  


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        context = request.form.get("context")
        num_cases = int(request.form.get("num_cases", 10))

        try:
            uploaded_file = request.files.get("file")
            if uploaded_file and uploaded_file.filename:
                file_path = os.path.join(upload_file_path, uploaded_file.filename)
                uploaded_file.save(file_path)
                generate_test_cases(context, num_cases, file_path, True)
                automate_python_testing(file_path, uploaded_file.filename)
                return save_test_results(context)
            else:
                return generate_test_cases(context, num_cases)
        except Exception as e:
            print("Error:", str(e))
            return jsonify({"error": str(e)})  
    return render_template("index.html")

@app.route("/download")
def download_file():
        return send_file(excel_file_path, as_attachment=True)

@app.route("/multi_download")
def multi_download():
    files = [excel_file_path, test_file_path, result_file_path]  
    return jsonify({"files": files})

@app.route("/fetch_file")
def fetch_file():
    file_path = request.args.get("path")
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(file_path, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=False)
